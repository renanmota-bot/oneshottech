import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from .models import Empresa, Usuario, ChamadoSuporte, MensagemChamado
from .forms import RegistroStaffForm
from eventos.models import Evento, Vaga, Candidatura, PresencaPagamento, AvisoEvento, PropostaComercial
from datetime import date, timedelta
import math
import re
import urllib.parse
import json
import urllib.request

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def validar_cpf(cpf_input):
    cpf = re.sub(r'\D', '', str(cpf_input))
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False
    for i in range(9, 11):
        val = sum(int(cpf[num]) * ((i + 1) - num) for num in range(0, i))
        digit = ((val * 10) % 11) % 10
        if int(cpf[i]) != digit:
            return False
    return True


def obter_lat_lng_endereco(endereco):
    try:
        end_clean = re.sub(r'CEP:\s*\d+', '', endereco).strip(' .-,')
        url = f"https://nominatim.openstreetmap.org/search?format=json&q={urllib.parse.quote(end_clean)}"
        req = urllib.request.Request(url, headers={'User-Agent': 'OneShotTech/1.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode())
            if data and len(data) > 0:
                return float(data[0]['lat']), float(data[0]['lon'])
    except Exception:
        pass
    return None, None


def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


def redirect_por_perfil(user):
    if user.perfil == 'SUPER_ADMIN' or user.is_superuser:
        return redirect('super_admin_dashboard')
    elif user.perfil == 'ADMIN':
        return redirect('admin_dashboard')
    else:
        return redirect('staff_dashboard')


def login_view(request):
    if request.user.is_authenticated:
        return redirect_por_perfil(request.user)

    if request.method == 'POST':
        usuario_input = request.POST.get('username')
        senha_input = request.POST.get('password')
        
        user = authenticate(request, username=usuario_input, password=senha_input)
        if user is not None:
            login(request, user)
            return redirect_por_perfil(user)
        else:
            messages.error(request, 'E-mail ou senha inválidos.')

    return render(request, 'core/login.html')


@csrf_exempt
def registro_staff_view(request):
    empresa_id = request.GET.get('empresa')
    empresa = Empresa.objects.filter(id=empresa_id).first() if empresa_id else None

    form = RegistroStaffForm()

    if request.method == 'POST':
        data = request.POST.copy()
        data['username'] = data.get('email', '')

        form = RegistroStaffForm(data, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['senha'])
            user.perfil = 'STAFF'
            if empresa:
                user.empresa = empresa
            user.save()
            
            login(request, user)
            messages.success(request, 'Bem-vindo! Cadastro realizado com sucesso.')
            return redirect('staff_dashboard')

    return render(request, 'core/registro_staff.html', {
        'form': form, 
        'empresa': empresa,
        'post_data': request.POST if request.method == 'POST' else {}
    })


def logout_view(request):
    logout(request)
    return redirect('login')


# ==========================================
# 1. SUPER ADMIN MASTER
# ==========================================
@login_required
def super_admin_dashboard(request):
    if not (request.user.perfil == 'SUPER_ADMIN' or request.user.is_superuser):
        return redirect('login')
        
    if request.method == 'POST':
        acao = request.POST.get('acao')

        if acao == 'criar_empresa':
            nome_emp = request.POST.get('nome') or request.POST.get('nome_empresa')
            cnpj_emp = request.POST.get('cnpj')
            valor_p = request.POST.get('valor_plano', 150.00)
            plano_p = request.POST.get('plano') or request.POST.get('plano_empresa', 'BASICO')
            whatsapp_emp = request.POST.get('whatsapp')
            email_admin = request.POST.get('email_admin')
            senha_admin = request.POST.get('senha_admin')

            if Empresa.objects.filter(cnpj=cnpj_emp).exists():
                messages.error(request, 'CNPJ/CPF já existente!')
            else:
                emp = Empresa.objects.create(
                    nome=nome_emp, cnpj=cnpj_emp, valor_plano=valor_p, 
                    plano=plano_p, whatsapp=whatsapp_emp, status='ATIVO'
                )
                if email_admin and senha_admin:
                    Usuario.objects.create_user(
                        username=email_admin, email=email_admin, password=senha_admin, 
                        first_name=f'Admin {nome_emp}', empresa=emp, perfil='ADMIN', is_staff=True
                    )
                messages.success(request, f'Empresa #{emp.id} - {emp.nome} criada com sucesso!')
            return redirect('super_admin_dashboard')

        elif acao == 'salvar_apis_empresa_master':
            emp = get_object_or_404(Empresa, id=request.POST.get('empresa_id'))
            emp.inscricao_municipal = request.POST.get('inscricao_municipal', emp.inscricao_municipal)
            emp.nfse_api_key = request.POST.get('nfse_api_key', emp.nfse_api_key)
            emp.whatsapp_api_instancia = request.POST.get('whatsapp_api_instancia', emp.whatsapp_api_instancia)
            emp.whatsapp_api_token = request.POST.get('whatsapp_api_token', emp.whatsapp_api_token)
            emp.save()
            messages.success(request, f'⚡ Chaves de API de NFS-e e WhatsApp da empresa {emp.nome} salvas com sucesso!')
            return redirect('super_admin_dashboard')

        elif acao == 'alterar_status_empresa':
            emp = get_object_or_404(Empresa, id=request.POST.get('empresa_id'))
            emp.status = request.POST.get('status', emp.status)
            novo_plano = request.POST.get('plano')
            if novo_plano:
                emp.plano = novo_plano
                if novo_plano == 'PREMIUM' and float(emp.valor_plano) <= 150.0:
                    emp.valor_plano = 299.00
                elif novo_plano == 'BASICO' and float(emp.valor_plano) >= 299.0:
                    emp.valor_plano = 150.00

            novo_val = request.POST.get('valor_plano')
            if novo_val: 
                emp.valor_plano = novo_val

            emp.save()
            messages.success(request, f'Empresa #{emp.id} atualizada! Plano: {emp.get_plano_display()} | Status: {emp.status}')
            return redirect('super_admin_dashboard')

        elif acao == 'alternar_bloqueio':
            emp = get_object_or_404(Empresa, id=request.POST.get('empresa_id'))
            if emp.status == 'BLOQUEADO':
                emp.status = 'ATIVO'
                messages.success(request, f'Empresa #{emp.id} DESBLOQUEADA!')
            else:
                emp.status = 'BLOQUEADO'
                messages.warning(request, f'Empresa #{emp.id} BLOQUEADA com sucesso!')
            emp.save()
            return redirect('super_admin_dashboard')

        elif acao == 'criar_usuario_global':
            nome_u = request.POST.get('nome')
            email_u = request.POST.get('email')
            senha_u = request.POST.get('senha')
            perfil_u = request.POST.get('perfil')
            emp_id_u = request.POST.get('empresa_id')
            emp_u = Empresa.objects.filter(id=emp_id_u).first() if emp_id_u else None

            if Usuario.objects.filter(username=email_u).exists():
                messages.error(request, 'E-mail já cadastrado no sistema!')
            else:
                Usuario.objects.create_user(username=email_u, email=email_u, password=senha_u, first_name=nome_u, perfil=perfil_u, empresa=emp_u)
                messages.success(request, f'Usuário {nome_u} criado com sucesso!')
            return redirect('super_admin_dashboard')

        elif acao == 'editar_usuario_global':
            usr = get_object_or_404(Usuario, id=request.POST.get('usuario_id'))
            usr.first_name = request.POST.get('nome', usr.first_name)
            usr.email = request.POST.get('email', usr.email)
            usr.username = request.POST.get('email', usr.username)
            usr.perfil = request.POST.get('perfil', usr.perfil)
            
            emp_id_u = request.POST.get('empresa_id')
            usr.empresa = Empresa.objects.filter(id=emp_id_u).first() if emp_id_u and emp_id_u != 'sem_empresa' else None

            senha_u = request.POST.get('senha')
            if senha_u and senha_u.strip():
                usr.set_password(senha_u.strip())

            usr.save()
            messages.success(request, f'Usuário #{usr.id} - {usr.first_name or usr.username} atualizado!')
            return redirect('super_admin_dashboard')

    busca = request.GET.get('busca', '').strip()
    filtro_perfil = request.GET.get('perfil', '').strip()
    filtro_empresa = request.GET.get('empresa_id', '').strip()

    usuarios_qs = Usuario.objects.all().select_related('empresa').order_by('-id')

    if busca:
        usuarios_qs = usuarios_qs.filter(first_name__icontains=busca) | usuarios_qs.filter(email__icontains=busca) | usuarios_qs.filter(cpf__icontains=busca)
    if filtro_perfil:
        usuarios_qs = usuarios_qs.filter(perfil=filtro_perfil)
    if filtro_empresa:
        if filtro_empresa == 'sem_empresa':
            usuarios_qs = usuarios_qs.filter(empresa__isnull=True)
        else:
            usuarios_qs = usuarios_qs.filter(empresa_id=filtro_empresa)

    usuarios_empresas = usuarios_qs.filter(perfil__in=['ADMIN', 'SUPER_ADMIN'])
    usuarios_staffs = usuarios_qs.filter(perfil='STAFF')

    empresas = Empresa.objects.all().order_by('-id')
    chamados_qs = ChamadoSuporte.objects.all().prefetch_related('mensagens', 'empresa').order_by('-id')

    total_mrr = sum([float(e.valor_plano) for e in empresas if e.status == 'ATIVO'])
    total_arr = total_mrr * 12

    context = {
        'empresas': empresas,
        'usuarios_empresas': usuarios_empresas,
        'usuarios_staffs': usuarios_staffs,
        'todos_usuarios': usuarios_qs,
        'chamados': chamados_qs,
        'total_empresas': empresas.count(),
        'total_empresas_ativas': empresas.filter(status='ATIVO').count(),
        'total_empresas_bloqueadas': empresas.filter(status='BLOQUEADO').count(),
        'total_staffs': Usuario.objects.filter(perfil='STAFF').count(),
        'total_eventos': Evento.objects.count(),
        'total_candidaturas': Candidatura.objects.count(),
        'total_checkins': PresencaPagamento.objects.filter(status_deslocamento='VALIDADO').count(),
        'chamados_abertos': ChamadoSuporte.objects.filter(status='ABERTO').count(),
        'total_mrr': total_mrr,
        'total_arr': total_arr,
        'busca': busca,
        'filtro_perfil': filtro_perfil,
        'filtro_empresa': filtro_empresa,
    }
    return render(request, 'core/super_admin.html', context)


@login_required
def ghost_login_view(request, user_id):
    if not (request.user.perfil == 'SUPER_ADMIN' or request.user.is_superuser):
        return redirect('login')
        
    target_user = get_object_or_404(Usuario, id=user_id)
    login(request, target_user, backend='django.contrib.auth.backends.ModelBackend')
    messages.info(request, f'👁️ Acesso Ghost ativado: Você está logado como "{target_user.first_name or target_user.username}"')
    return redirect_por_perfil(target_user)


# ==========================================
# 2. ADMIN PRODUTORA
# ==========================================
@login_required
def admin_dashboard(request):
    if not (request.user.perfil == 'ADMIN' or request.user.is_superuser):
        return redirect('login')

    empresa = request.user.empresa or Empresa.objects.first()

    # BLOQUEIO TOTAL DE CONTA SUSPENSA
    if empresa and empresa.status == 'BLOQUEADO' and not request.user.is_superuser:
        return render(request, 'core/conta_bloqueada.html', {'empresa': empresa})

    hoje = date.today()

    if request.method == 'POST':
        acao = request.POST.get('acao')

        if acao == 'criar_proposta':
            if not empresa.is_premium:
                messages.error(request, '🔒 Recurso de Propostas Comerciais exclusivo do Plano Premium.')
                return redirect('admin_dashboard')

            evento_id = request.POST.get('evento_id')
            evento = Evento.objects.filter(id=evento_id, empresa=empresa).first() if evento_id else None
            val_total = str(request.POST.get('valor_total', '0')).replace(',', '.')

            PropostaComercial.objects.create(
                empresa=empresa,
                evento=evento,
                cliente_nome=request.POST.get('cliente_nome'),
                cliente_cnpj_cpf=request.POST.get('cliente_cnpj_cpf'),
                cliente_email=request.POST.get('cliente_email'),
                cliente_endereco=request.POST.get('cliente_endereco'),
                valor_total=val_total,
                descricao_servicos=request.POST.get('descricao_servicos'),
                status=request.POST.get('status', 'RASCUNHO')
            )
            messages.success(request, '📑 Proposta Comercial cadastrada com sucesso!')
            return redirect('admin_dashboard')

        elif acao == 'alterar_status_proposta':
            proposta = get_object_or_404(PropostaComercial, id=request.POST.get('proposta_id'), empresa=empresa)
            proposta.status = request.POST.get('novo_status')
            proposta.save()
            messages.success(request, f'Status da Proposta #{proposta.id} atualizado para {proposta.get_status_display()}!')
            return redirect('admin_dashboard')

        elif acao == 'emitir_nfse_proposta':
            if not empresa.is_premium:
                messages.error(request, '🔒 Emissão de NFS-e exclusiva do Plano Premium.')
                return redirect('admin_dashboard')

            proposta = get_object_or_404(PropostaComercial, id=request.POST.get('proposta_id'), empresa=empresa)
            proposta.nfse_status = 'EMITIDA'
            proposta.nfse_numero = f"NFS-{proposta.id:06d}"
            proposta.save()
            messages.success(request, f'📄 Nota Fiscal Eletrônica #{proposta.nfse_numero} emitida com sucesso para o cliente {proposta.cliente_nome}!')
            return redirect('admin_dashboard')

        elif acao == 'criar_evento':
            data_inicio_input = request.POST.get('data_inicio')
            data_termino_input = request.POST.get('data_termino')
            orcamento_prev = request.POST.get('orcamento_previsto', 0.00)

            if not data_termino_input:
                messages.error(request, '❌ A Data de Término é obrigatória.')
                return redirect('admin_dashboard')

            if data_termino_input < data_inicio_input:
                messages.error(request, '❌ A Data de Término não pode ser anterior à Data de Início.')
                return redirect('admin_dashboard')

            cep = request.POST.get('cep', '')
            rua = request.POST.get('rua', '')
            numero = request.POST.get('numero', '')
            bairro = request.POST.get('bairro', '')
            cidade = request.POST.get('cidade', '')
            uf = request.POST.get('uf', '')
            
            if cep or rua:
                endereco_local = f"{rua}, {numero} - {bairro}, {cidade} - {uf}. CEP: {cep}"
            else:
                endereco_local = request.POST.get('local', 'Local a definir')

            lat_auto, lng_auto = obter_lat_lng_endereco(endereco_local)

            evento_data = {
                'empresa': empresa,
                'nome': request.POST.get('nome'),
                'local': endereco_local,
                'latitude': lat_auto,
                'longitude': lng_auto,
                'data_inicio': data_inicio_input,
                'orcamento_previsto': str(orcamento_prev).replace(',', '.'),
            }
            if hasattr(Evento, 'data_termino'):
                evento_data['data_termino'] = data_termino_input

            Evento.objects.create(**evento_data)
            messages.success(request, 'Evento criado com sucesso!')
            return redirect('admin_dashboard')

        elif acao == 'criar_vaga':
            evento = get_object_or_404(Evento, id=request.POST.get('evento_id'), empresa=empresa)
            val_raw = str(request.POST.get('valor_diaria', '0')).replace(',', '.')
            dress_code_input = request.POST.get('dress_code', '')

            vaga_data = {
                'evento': evento,
                'funcao': request.POST.get('funcao'),
                'valor_diaria': val_raw,
                'quantidade': request.POST.get('quantidade', 1),
                'prazo_pagamento_dias': request.POST.get('prazo_pagamento_dias', 0)
            }
            if hasattr(Vaga, 'dress_code') and dress_code_input:
                vaga_data['dress_code'] = dress_code_input

            Vaga.objects.create(**vaga_data)
            messages.success(request, 'Vaga criada com sucesso!')
            return redirect('admin_dashboard')

        elif acao == 'postar_aviso':
            evento = get_object_or_404(Evento, id=request.POST.get('evento_id'), empresa=empresa)
            titulo = request.POST.get('titulo')
            mensagem_texto = request.POST.get('mensagem')
            
            AvisoEvento.objects.create(evento=evento, titulo=titulo, mensagem=mensagem_texto)
            messages.success(request, f'📢 Aviso publicado no mural do evento {evento.nome}!')
            return redirect('admin_dashboard')

        elif acao == 'avaliar_staff':
            pres = get_object_or_404(PresencaPagamento, id=request.POST.get('presenca_id'), candidatura__vaga__evento__empresa=empresa)
            nota = int(request.POST.get('nota_desempenho', 5))
            comentario = request.POST.get('comentario_desempenho', '')

            pres.nota_desempenho = nota
            pres.comentario_desempenho = comentario
            pres.save()

            user_staff = pres.candidatura.usuario
            avaliacoes = PresencaPagamento.objects.filter(candidatura__usuario=user_staff, nota_desempenho__isnull=False)
            if avaliacoes.exists():
                media = sum(a.nota_desempenho for a in avaliacoes) / avaliacoes.count()
                user_staff.nota_media = round(media, 1)
                user_staff.save()

            messages.success(request, f'⭐ Avaliação de {user_staff.first_name or user_staff.username} salva com sucesso!')
            return redirect('admin_dashboard')

        elif acao == 'alterar_candidatura':
            cand = get_object_or_404(Candidatura, id=request.POST.get('candidatura_id'), vaga__evento__empresa=empresa)
            novo_status = request.POST.get('novo_status')
            cand.status = novo_status
            cand.save()

            vaga = cand.vaga
            total_aprovados = Candidatura.objects.filter(vaga=vaga, status='APROVADO').count()

            if novo_status == 'APROVADO':
                PresencaPagamento.objects.get_or_create(candidatura=cand)
                if total_aprovados >= vaga.quantidade:
                    vaga.status = 'PREENCHIDA'
                    vaga.save()
            else:
                if total_aprovados < vaga.quantidade and vaga.status == 'PREENCHIDA':
                    vaga.status = 'ABERTA'
                    vaga.save()

            messages.success(request, f'Status atualizado para {cand.status}.')
            return redirect('admin_dashboard')

        elif acao == 'validar_checkin':
            pres = get_object_or_404(PresencaPagamento, id=request.POST.get('presenca_id'), candidatura__vaga__evento__empresa=empresa)
            pres.status_deslocamento = 'VALIDADO'
            pres.save()
            messages.success(request, 'Check-in validado! Staff enviado para Pagamentos.')
            return redirect('admin_dashboard')

        elif acao == 'marcar_falta':
            pres = get_object_or_404(PresencaPagamento, id=request.POST.get('presenca_id'), candidatura__vaga__evento__empresa=empresa)
            pres.status_deslocamento = 'FALTOU'
            pres.status_pagamento = 'CANCELADO'
            pres.save()

            vaga = pres.candidatura.vaga
            total_aprovados = Candidatura.objects.filter(vaga=vaga, status='APROVADO').exclude(candidatura__presenca_pagamento__status_deslocamento='FALTOU').count()
            if total_aprovados < vaga.quantidade and vaga.status == 'PREENCHIDA':
                vaga.status = 'ABERTA'
                vaga.save()

            messages.error(request, 'Falta registrada. Vaga reaberta no sistema.')
            return redirect('admin_dashboard')

        elif acao == 'marcar_pago':
            pres = get_object_or_404(PresencaPagamento, id=request.POST.get('presenca_id'), candidatura__vaga__evento__empresa=empresa)
            pres.status_pagamento = 'PAGO'
            
            evento = pres.candidatura.vaga.evento
            dias_totais = 1
            if hasattr(evento, 'data_termino') and evento.data_termino and evento.data_inicio:
                dias_totais = (evento.data_termino - evento.data_inicio).days + 1
                
            dias = pres.dias_presentes if getattr(pres, 'dias_presentes', 0) > 0 else dias_totais
            pres.valor_pago = dias * pres.candidatura.vaga.valor_diaria
            pres.save()
            
            messages.success(request, 'Pagamento baixado! O staff foi notificado para confirmar o recebimento.')
            return redirect('admin_dashboard')

        elif acao == 'pagar_em_massa':
            presenca_ids = request.POST.getlist('presenca_ids')
            if not presenca_ids:
                messages.warning(request, 'Nenhum pagamento foi selecionado.')
                return redirect('admin_dashboard')
                
            for p_id in presenca_ids:
                pres = PresencaPagamento.objects.filter(id=p_id, candidatura__vaga__evento__empresa=empresa).first()
                if pres and pres.status_pagamento not in ['PAGO', 'CONFIRMADO']:
                    pres.status_pagamento = 'PAGO'
                    
                    evento = pres.candidatura.vaga.evento
                    dias_totais = 1
                    if hasattr(evento, 'data_termino') and evento.data_termino and evento.data_inicio:
                        dias_totais = (evento.data_termino - evento.data_inicio).days + 1
                        
                    dias = pres.dias_presentes if getattr(pres, 'dias_presentes', 0) > 0 else dias_totais
                    pres.valor_pago = dias * pres.candidatura.vaga.valor_diaria
                    pres.save()
                    
            messages.success(request, f'{len(presenca_ids)} pagamentos realizados com sucesso e notificados aos staffs!')
            return redirect('admin_dashboard')

        elif acao == 'editar_staff_empresa':
            usr = get_object_or_404(Usuario, id=request.POST.get('usuario_id'), empresa=empresa)
            usr.first_name = request.POST.get('first_name', usr.first_name)
            usr.last_name = request.POST.get('last_name', usr.last_name)
            usr.email = request.POST.get('email', usr.email)
            usr.whatsapp = request.POST.get('whatsapp', usr.whatsapp)
            
            novo_cpf = request.POST.get('cpf', usr.cpf)
            if novo_cpf and novo_cpf != usr.cpf:
                if validar_cpf(novo_cpf):
                    usr.cpf = novo_cpf
                else:
                    messages.error(request, f'❌ O CPF informado ({novo_cpf}) é inválido.')
                    return redirect('admin_dashboard')
            
            usr.rg = request.POST.get('rg', usr.rg)
            usr.tipo_chave_pix = request.POST.get('tipo_chave_pix', usr.tipo_chave_pix)
            usr.chave_pix = request.POST.get('chave_pix', usr.chave_pix)
            usr.save()
            messages.success(request, f'✅ Cadastro do colaborador atualizado com sucesso!')
            return redirect('admin_dashboard')

    eventos = Evento.objects.filter(empresa=empresa).order_by('-data_inicio') if empresa else []
    
    eventos_ativos = [ev for ev in eventos if (getattr(ev, 'data_termino', ev.data_inicio) or ev.data_inicio) >= hoje]
    eventos_concluidos = [ev for ev in eventos if (getattr(ev, 'data_termino', ev.data_inicio) or ev.data_inicio) < hoje]

    vagas = Vaga.objects.filter(evento__empresa=empresa).order_by('-id') if empresa else []
    
    operacao = PresencaPagamento.objects.filter(
        candidatura__vaga__evento__empresa=empresa,
        candidatura__status='APROVADO'
    ).exclude(status_deslocamento__in=['VALIDADO', 'FALTOU']).order_by('-id') if empresa else []

    checkins_pendentes_count = operacao.filter(status_deslocamento='CHECKIN_REALIZADO').count()

    financeiro = PresencaPagamento.objects.filter(
        candidatura__vaga__evento__empresa=empresa,
        candidatura__status='APROVADO',
        status_deslocamento='VALIDADO'
    ).order_by('-id') if empresa else []

    propostas = PropostaComercial.objects.filter(empresa=empresa).order_by('-id') if empresa else []
    equipe_staff = Usuario.objects.filter(empresa=empresa, perfil='STAFF').order_by('-id') if empresa else []

    total_eventos = eventos.count()
    total_staffs = operacao.count() + financeiro.count()
    
    total_caches_previstos = sum([f.candidatura.vaga.valor_diaria for f in financeiro if f.status_pagamento not in ['PAGO', 'CONFIRMADO']])
    total_caches_pagos = sum([f.candidatura.vaga.valor_diaria for f in financeiro if f.status_pagamento in ['PAGO', 'CONFIRMADO']])

    link_convite = request.build_absolute_uri(f"/registro/staff/?empresa={empresa.id}") if empresa else ""

    context = {
        'empresa': empresa,
        'eventos_ativos': eventos_ativos,
        'eventos_concluidos': eventos_concluidos,
        'vagas': vagas,
        'operacao': operacao,
        'checkins_pendentes_count': checkins_pendentes_count,
        'financeiro': financeiro,
        'propostas': propostas,
        'equipe_staff': equipe_staff,
        'total_eventos': total_eventos,
        'total_staffs': total_staffs,
        'total_caches_previstos': total_caches_previstos,
        'total_caches_pagos': total_caches_pagos,
        'link_convite': link_convite,
    }
    return render(request, 'core/admin_dashboard.html', context)


# ==========================================
# EXPORTAÇÃO LOTE PIX CSV (PREMIUM)
# ==========================================
@login_required
def exportar_lote_pix_csv(request):
    if not (request.user.perfil in ['ADMIN', 'SUPER_ADMIN'] or request.user.is_superuser):
        return redirect('login')

    empresa = request.user.empresa or Empresa.objects.first()
    if not empresa or not empresa.is_premium:
        messages.error(request, '🔒 Exportação de Lote PIX em CSV exclusiva para o Plano Premium.')
        return redirect('admin_dashboard')

    financeiro = PresencaPagamento.objects.filter(
        candidatura__vaga__evento__empresa=empresa,
        candidatura__status='APROVADO',
        status_deslocamento='VALIDADO',
        status_pagamento='PENDENTE'
    ).select_related('candidatura__usuario', 'candidatura__vaga__evento')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="lote_pix_pagamentos_{empresa.id}.csv"'
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nome do Beneficiario', 'CPF Beneficiario', 'Tipo de Chave PIX', 'Chave PIX', 'Valor (R$)', 'Evento', 'Funcao'])

    for p in financeiro:
        usr = p.candidatura.usuario
        val = p.candidatura.vaga.valor_diaria
        writer.writerow([
            usr.get_full_name() or usr.username,
            usr.cpf or 'N/A',
            usr.tipo_chave_pix or 'CPF/CNPJ',
            usr.chave_pix or 'N/A',
            f"{val:.2f}".replace('.', ','),
            p.candidatura.vaga.evento.nome,
            p.candidatura.vaga.funcao
        ])

    return response


# ==========================================
# 3. STAFF
# ==========================================
@login_required
def staff_dashboard(request):
    if request.user.perfil not in ['STAFF', 'SUPER_ADMIN'] and not request.user.is_superuser:
        return redirect('login')

    hoje = date.today()

    minhas_candidaturas = Candidatura.objects.filter(usuario=request.user).select_related('vaga__evento__empresa', 'presenca_pagamento').order_by('-id')
    vagas_candidatadas_ids = list(minhas_candidaturas.values_list('vaga_id', flat=True))
    vagas_aprovadas_todas = minhas_candidaturas.filter(status='APROVADO')

    vagas_aprovadas = []
    eventos_finalizados_staff = []
    datas_bloqueadas = set()

    for cand in vagas_aprovadas_todas:
        ev = cand.vaga.evento
        dt_inicio = ev.data_inicio
        dt_fim = getattr(ev, 'data_termino', None) or dt_inicio
        pres = getattr(cand, 'presenca_pagamento', None)
        st_pagamento = pres.status_pagamento if pres else 'PENDENTE'

        if dt_fim < hoje or st_pagamento in ['PAGO', 'CONFIRMADO']:
            eventos_finalizados_staff.append(cand)
        else:
            vagas_aprovadas.append(cand)

            cur_dt = dt_inicio
            while cur_dt <= dt_fim:
                datas_bloqueadas.add(cur_dt)
                cur_dt += timedelta(days=1)

    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'candidatar':
            vaga = get_object_or_404(Vaga, id=request.POST.get('vaga_id'))
            ev_vaga = vaga.evento

            dt_inicio_vaga = ev_vaga.data_inicio
            dt_fim_vaga = getattr(ev_vaga, 'data_termino', None) or dt_inicio_vaga
            
            datas_vaga = set()
            cur_dt = dt_inicio_vaga
            while cur_dt <= dt_fim_vaga:
                datas_vaga.add(cur_dt)
                cur_dt += timedelta(days=1)

            if datas_bloqueadas.intersection(datas_vaga):
                messages.error(request, '❌ Você já está aprovado em outro evento nesta mesma data e não pode aceitar novas vagas neste período.')
                return redirect('staff_dashboard')

            aprovados_count = Candidatura.objects.filter(vaga=vaga, status='APROVADO').count()
            if aprovados_count >= vaga.quantidade or vaga.status == 'PREENCHIDA':
                messages.error(request, '❌ Esta vaga já foi preenchida.')
                return redirect('staff_dashboard')

            cand, created = Candidatura.objects.get_or_create(vaga=vaga, usuario=request.user)
            if created:
                cand.aceitou_termo = True
                cand.data_aceite_termo = timezone.now()
                cand.save()
                messages.success(request, f'Inscrição efetuada com sucesso para {vaga.funcao}! Termo e contrato aceitos.')
            else:
                messages.warning(request, 'Você já se candidatou a esta vaga.')
            return redirect('staff_dashboard')

        elif acao == 'a_caminho':
            cand_id = request.POST.get('candidatura_id')
            cand = get_object_or_404(Candidatura, id=cand_id, usuario=request.user, status='APROVADO')
            pres, _ = PresencaPagamento.objects.get_or_create(candidatura=cand)
            pres.status_deslocamento = 'A_CAMINHO'
            pres.save()
            messages.success(request, 'Aviso enviado! A produtora já sabe que você está a caminho 🚗')
            return redirect('staff_dashboard')

        elif acao == 'checkin_gps':
            cand_id = request.POST.get('candidatura_id')
            lat_user = request.POST.get('lat_user')
            lng_user = request.POST.get('lng_user')
            
            if not lat_user or not lng_user or float(lat_user) == 0:
                messages.error(request, '❌ Não foi possível obter a sua localização. Verifique se o GPS do celular está ativo e se você concedeu permissão de acesso ao navegador.')
                return redirect('staff_dashboard')

            lat_user = float(lat_user)
            lng_user = float(lng_user)
            
            cand = get_object_or_404(Candidatura, id=cand_id, usuario=request.user, status='APROVADO')
            evento = cand.vaga.evento
            
            if not evento.latitude or not evento.longitude:
                lat_ev, lng_ev = obter_lat_lng_endereco(evento.local)
                if lat_ev and lng_ev:
                    evento.latitude, evento.longitude = lat_ev, lng_ev
                    evento.save()
                else:
                    evento.latitude, evento.longitude = lat_user, lng_user
                    evento.save()

            distancia = haversine_distance(lat_user, lng_user, float(evento.latitude), float(evento.longitude))
            
            if distancia > 50:
                messages.error(request, f'❌ Check-in recusado! Você está a {int(distancia)} metros do evento. É necessário estar a menos de 50 metros do local para confirmar presença.')
                return redirect('staff_dashboard')

            pres, _ = PresencaPagamento.objects.get_or_create(candidatura=cand)
            pres.status_deslocamento = 'CHECKIN_REALIZADO'
            pres.lat_checkin = lat_user
            pres.lng_checkin = lng_user

            if 'foto_checkin' in request.FILES:
                pres.foto_checkin = request.FILES['foto_checkin']

            pres.save()
            messages.success(request, f'📍 Check-in com Selfie realizado com sucesso! Você está a {int(distancia)}m do evento.')
            return redirect('staff_dashboard')

        elif acao == 'confirmar_recebimento':
            pres = get_object_or_404(PresencaPagamento, id=request.POST.get('presenca_id'), candidatura__usuario=request.user)
            if pres.status_pagamento == 'PAGO':
                pres.status_pagamento = 'CONFIRMADO'
                pres.save()
                messages.success(request, '✅ Recebimento confirmado! Muito obrigado.')
            return redirect('staff_dashboard')

        elif acao == 'preencher_ficha':
            user = request.user
            if user.cpf and user.rg:
                messages.warning(request, '🔒 Sua ficha cadastral já está salva e bloqueada. Para alterar, solicite à produtora.')
                return redirect('staff_dashboard')

            nome = request.POST.get('first_name')
            cpf = request.POST.get('cpf')
            rg = request.POST.get('rg')

            if not validar_cpf(cpf):
                messages.error(request, '❌ CPF Inválido! Verifique os números e tente novamente.')
                return redirect('staff_dashboard')

            if not rg or len(rg.strip()) < 5:
                messages.error(request, '❌ Documento RG Inválido!')
                return redirect('staff_dashboard')

            user.first_name = nome
            user.cpf = cpf
            user.rg = rg
            user.save()
            
            messages.success(request, '📋 Ficha Cadastral confirmada e salva com sucesso!')
            return redirect('staff_dashboard')

        elif acao == 'atualizar_perfil':
            user = request.user
            
            if not (user.cpf and user.rg):
                novo_cpf = request.POST.get('cpf', user.cpf)
                if novo_cpf and not validar_cpf(novo_cpf):
                    messages.error(request, '❌ Erro ao salvar: O CPF informado é inválido!')
                    return redirect('staff_dashboard')
                user.cpf = novo_cpf
                user.rg = request.POST.get('rg', user.rg)

            user.first_name = request.POST.get('first_name', user.first_name)
            user.last_name = request.POST.get('last_name', user.last_name)
            user.email = request.POST.get('email', user.email)
            user.whatsapp = request.POST.get('whatsapp', user.whatsapp)
            user.tipo_chave_pix = request.POST.get('tipo_chave_pix', user.tipo_chave_pix)
            user.chave_pix = request.POST.get('chave_pix', user.chave_pix)

            senha = request.POST.get('senha')
            if senha and senha.strip():
                user.set_password(senha.strip())

            if 'foto' in request.FILES:
                user.foto = request.FILES['foto']

            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, '✅ Perfil atualizado!')
            return redirect('staff_dashboard')

    vagas_disponiveis_qs = Vaga.objects.filter(status='ABERTA').select_related('evento__empresa').order_by('-id')

    vagas_disponiveis = []
    for v in vagas_disponiveis_qs:
        dt_fim_vaga = getattr(v.evento, 'data_termino', None) or v.evento.data_inicio
        if dt_fim_vaga < hoje:
            continue

        aprovados_count = Candidatura.objects.filter(vaga=v, status='APROVADO').count()
        if aprovados_count < v.quantidade:
            v.vagas_restantes = v.quantidade - aprovados_count
            
            dt_inicio_vaga = v.evento.data_inicio
            datas_vaga = set()
            cur_dt = dt_inicio_vaga
            while cur_dt <= dt_fim_vaga:
                datas_vaga.add(cur_dt)
                cur_dt += timedelta(days=1)

            v.tem_conflito_data = bool(datas_bloqueadas.intersection(datas_vaga))
            vagas_disponiveis.append(v)

    extrato_pagamentos = PresencaPagamento.objects.filter(
        candidatura__usuario=request.user,
        candidatura__status='APROVADO'
    ).select_related('candidatura__vaga__evento__empresa').order_by('-id')

    pagamentos_pendentes_confirmacao = sum(1 for p in extrato_pagamentos if p.status_pagamento == 'PAGO')

    tot_recebido = sum(p.candidatura.vaga.valor_diaria for p in extrato_pagamentos if p.status_pagamento in ['PAGO', 'CONFIRMADO'])
    tot_a_receber = sum(p.candidatura.vaga.valor_diaria for p in extrato_pagamentos if p.status_pagamento not in ['PAGO', 'CONFIRMADO'])

    context = {
        'vagas_disponiveis': vagas_disponiveis,
        'vagas_candidatadas_ids': vagas_candidatadas_ids,
        'vagas_aprovadas': vagas_aprovadas,
        'eventos_finalizados_staff': eventos_finalizados_staff,
        'extrato_pagamentos': extrato_pagamentos,
        'pagamentos_pendentes_confirmacao': pagamentos_pendentes_confirmacao,
        'tot_recebido': tot_recebido,
        'tot_a_receber': tot_a_receber,
        'ficha_bloqueada': bool(request.user.cpf and request.user.rg)
    }
    return render(request, 'core/staff_dashboard.html', context)


# ==========================================
# EXPORTAÇÕES DE RELATÓRIOS
# ==========================================
@login_required
def exportar_caches_excel(request):
    if not (request.user.perfil in ['ADMIN', 'SUPER_ADMIN'] or request.user.is_superuser):
        return redirect('login')

    empresa = request.user.empresa or Empresa.objects.first()
    pagamentos = Candidatura.objects.filter(vaga__evento__empresa=empresa, status='APROVADO').select_related('presenca_pagamento', 'usuario', 'vaga__evento').order_by('-id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aprovados - Relatorio Caches"

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ["ID Staff", "Nome", "Sobrenome", "CPF", "RG", "WhatsApp", "Gênero", "Camiseta", "Calçado", "Tipo Pix", "Chave Pix", "Evento", "Função", "Valor Diária (R$)", "Prazo Pagamento (Dias)", "Status Pagamento"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    for item in pagamentos:
        u = item.usuario
        pres = getattr(item, 'presenca_pagamento', None)
        st_p = pres.status_pagamento if pres else 'PENDENTE'
        v_diaria = float(item.vaga.valor_diaria)

        row = [u.id, u.first_name or u.username, u.last_name or "", u.cpf or "N/A", u.rg or "N/A", u.whatsapp or "N/A", u.get_genero_display() if u.genero else "N/A", u.tamanho_camiseta or "N/A", u.tamanho_calcado or "N/A", u.tipo_chave_pix or "N/A", u.chave_pix or "N/A", item.vaga.evento.nome, item.vaga.funcao, v_diaria, item.vaga.prazo_pagamento_dias, st_p]
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_aprovados_{empresa.id if empresa else 1}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_caches_pdf(request):
    if not (request.user.perfil in ['ADMIN', 'SUPER_ADMIN'] or request.user.is_superuser):
        return redirect('login')

    empresa = request.user.empresa or Empresa.objects.first()
    pagamentos = Candidatura.objects.filter(vaga__evento__empresa=empresa, status='APROVADO').select_related('presenca_pagamento', 'usuario', 'vaga__evento').order_by('-id')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="folha_caches_{empresa.id if empresa else 1}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor('#0F172A'), spaceAfter=6)
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#64748B'), spaceAfter=15)

    elements.append(Paragraph(f"One Shot Tech — Folha de Pagamento de Cachês", title_style))
    elements.append(Paragraph(f"Produtora: {empresa.nome if empresa else 'Global'} | Emitido para o Financeiro", subtitle_style))

    data = [["Staff", "CPF", "Pix", "Evento / Função", "Diária", "Prazo", "Status"]]

    for item in pagamentos:
        pres = getattr(item, 'presenca_pagamento', None)
        st_p = pres.status_pagamento if pres else 'PENDENTE'
        v_diaria = float(item.vaga.valor_diaria)

        data.append([
            item.usuario.get_full_name() or item.usuario.username,
            item.usuario.cpf or "N/A",
            item.usuario.chave_pix or "N/A",
            f"{item.vaga.evento.nome}\n({item.vaga.funcao})",
            f"R$ {v_diaria:.2f}",
            f"{item.vaga.prazo_pagamento_dias} dias",
            st_p
        ])

    table = Table(data, colWidths=[110, 80, 85, 140, 55, 45, 55])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_ficha_staff_pdf(request, user_id):
    if not (request.user.perfil in ['ADMIN', 'SUPER_ADMIN'] or request.user.is_superuser):
        return redirect('login')

    staff_user = get_object_or_404(Usuario, id=user_id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ficha_staff_{staff_user.id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#0F172A'), spaceAfter=4)
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#64748B'), spaceAfter=20)
    section_style = ParagraphStyle('SectionStyle', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor('#0284C7'), spaceBefore=10, spaceAfter=8)

    elements.append(Paragraph(f"Ficha Cadastral do Colaborador", title_style))
    elements.append(Paragraph(f"One Shot Tech | Plataforma de Casting & Eventos", subtitle_style))

    dados_pessoais = [
        ["Nome Completo:", staff_user.get_full_name() or staff_user.username],
        ["CPF:", staff_user.cpf or "Não informado"],
        ["RG:", staff_user.rg or "Não informado"],
        ["Gênero:", staff_user.get_genero_display() if staff_user.genero else "Não informado"],
        ["WhatsApp / Contato:", staff_user.whatsapp or "Não informado"],
        ["E-mail:", staff_user.email],
        ["Produtora Vinculada:", staff_user.empresa.nome if staff_user.empresa else "Global"],
    ]

    t_pessoais = Table(dados_pessoais, colWidths=[140, 380])
    t_pessoais.setStyle(TableStyle([
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#0F172A')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))

    elements.append(Paragraph("1. Dados Pessoais & Contatos", section_style))
    elements.append(t_pessoais)
    elements.append(Spacer(1, 15))

    dados_operacionais = [
        ["Tamanho de Camiseta:", staff_user.tamanho_camiseta or "Não informado"],
        ["Número de Calçado:", staff_user.tamanho_calcado or "Não informado"],
        ["Tipo de Chave Pix:", staff_user.tipo_chave_pix or "Não informado"],
        ["Chave Pix:", staff_user.chave_pix or "Não informado"],
        ["Nota Média de Avaliação:", f"⭐ {staff_user.nota_media} / 5.0"],
    ]

    t_operacionais = Table(dados_operacionais, colWidths=[140, 380])
    t_operacionais.setStyle(TableStyle([
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#0F172A')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
    ]))

    elements.append(Paragraph("2. Dress Code & Dados Financeiros", section_style))
    elements.append(t_operacionais)

    doc.build(elements)
    return response


@login_required
def exportar_extrato_staff_pdf(request):
    if request.user.perfil not in ['STAFF', 'SUPER_ADMIN'] and not request.user.is_superuser:
        return redirect('login')

    user = request.user
    candidaturas = Candidatura.objects.filter(usuario=user, status='APROVADO').select_related('vaga__evento', 'presenca_pagamento').order_by('-id')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="extrato_caches_{user.username}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, textColor=colors.HexColor('#0F172A'), spaceAfter=4)
    subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#64748B'), spaceAfter=15)

    elements.append(Paragraph(f"Extrato Financeiro de Cachês", title_style))
    elements.append(Paragraph(f"Colaborador: {user.get_full_name() or user.username} | CPF: {user.cpf or 'N/A'}", subtitle_style))

    data = [["Evento", "Função", "Diária (R$)", "Dias Presentes", "Total Calculado (R$)", "Status Pagamento"]]

    tot_geral = 0.0
    for c in candidaturas:
        pres = getattr(c, 'presenca_pagamento', None)
        dias = pres.dias_presentes if pres else 0
        st_p = pres.status_pagamento if pres else 'PENDENTE'
        v_diaria = float(c.vaga.valor_diaria)
        v_total = v_diaria * dias
        tot_geral += v_total

        data.append([
            c.vaga.evento.nome,
            c.vaga.funcao,
            f"R$ {v_diaria:.2f}",
            str(dias),
            f"R$ {v_total:.2f}",
            st_p
        ])

    data.append(["TOTAL GERAL", "", "", "", f"R$ {tot_geral:.2f}", ""])

    table = Table(data, colWidths=[130, 100, 70, 75, 95, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F1F5F9')),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required
def exportar_evento_excel(request, evento_id):
    if not (request.user.perfil in ['ADMIN', 'SUPER_ADMIN'] or request.user.is_superuser):
        return redirect('login')

    empresa = request.user.empresa or Empresa.objects.first()
    evento = get_object_or_404(Evento, id=evento_id, empresa=empresa)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Evento - {evento.nome[:20]}"

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    headers = ["ID Staff", "Nome", "CPF", "WhatsApp", "Função", "Valor Diária (R$)", "Status Pagamento", "Total Pago (R$)"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    candidaturas = Candidatura.objects.filter(vaga__evento=evento, status='APROVADO').select_related('usuario', 'vaga', 'presenca_pagamento')

    for item in candidaturas:
        u = item.usuario
        pres = getattr(item, 'presenca_pagamento', None)
        st_p = pres.status_pagamento if pres else 'PENDENTE'
        v_diaria = float(item.vaga.valor_diaria)
        v_pago = float(pres.valor_pago) if pres and pres.valor_pago else (v_diaria if st_p in ['PAGO', 'CONFIRMADO'] else 0.0)

        row = [u.id, u.get_full_name() or u.username, u.cpf or "N/A", u.whatsapp or "N/A", item.vaga.funcao, v_diaria, st_p, v_pago]
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_evento_{evento.id}.xlsx"'
    wb.save(response)
    return response